from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
from pathlib import Path

IDR_FORMAT = '#,##0'

def number_to_terbilang(number: float) -> str:
    """Convert number to Indonesian words (terbilang)"""
    if number == 0:
        return "nol rupiah"

    units = ["", "satu", "dua", "tiga", "empat", "lima", "enam", "tujuh", "delapan", "sembilan"]
    teens = ["sepuluh", "sebelas", "dua belas", "tiga belas", "empat belas", "lima belas", "enam belas", "tujuh belas", "delapan belas", "sembilan belas"]
    tens = ["", "", "dua puluh", "tiga puluh", "empat puluh", "lima puluh", "enam puluh", "tujuh puluh", "delapan puluh", "sembilan puluh"]

    def convert_chunk(n):
        if n < 10:
            return units[n]
        elif n < 20:
            return teens[n - 10]
        elif n < 100:
            return tens[n // 10] + (" " + units[n % 10] if n % 10 != 0 else "")
        elif n < 1000:
            return units[n // 100] + " ratus" + (" " + convert_chunk(n % 100) if n % 100 != 0 else "")
        elif n < 1000000:
            return convert_chunk(n // 1000) + " ribu" + (" " + convert_chunk(n % 1000) if n % 1000 != 0 else "")
        elif n < 1000000000:
            return convert_chunk(n // 1000000) + " juta" + (" " + convert_chunk(n % 1000000) if n % 1000000 != 0 else "")
        else:
            # Billions+
            billions = n // 1000000000
            remainder = n % 1000000000
            result = convert_chunk(billions) + " miliar"
            if remainder > 0:
                result += " " + convert_chunk(remainder)
            return result

    number_int = int(number)
    return convert_chunk(number_int) + " rupiah"


class ExcelService:
    def create_workbook(self) -> Workbook:
        return Workbook()

    def add_items_table(self, wb: Workbook, items: List[Dict], sheet_name: str = "RAB"):
        ws = wb.active
        ws.title = sheet_name

        # Headers
        headers = ["NO", "URAIAN", "VOLUME", "SATUAN", "HARGA SATUAN (IDR)", "JUMLAH HARGA (IDR)"]
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.border = thin_border

        # Data rows
        for idx, item in enumerate(items, start=1):
            row = idx + 1
            ws[f"A{row}"] = idx
            ws[f"B{row}"] = item.get("uraian", "")
            ws[f"C{row}"] = float(item.get("volume", 0))
            ws[f"D{row}"] = item.get("satuan", "")
            ws[f"E{row}"] = float(item.get("harga_satuan", 0))
            ws[f"F{row}"] = f"=C{row}*E{row}"

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            # Format number columns as IDR
            ws.cell(row=row, column=5).number_format = IDR_FORMAT  # HARGA SATUAN
            ws.cell(row=row, column=6).number_format = IDR_FORMAT  # JUMLAH HARGA

        last_data_row = len(items) + 1

        # Total row
        total_row = last_data_row + 1
        ws[f"E{total_row}"] = "Total"
        ws[f"E{total_row}"].font = Font(bold=True)
        ws[f"E{total_row}"].alignment = Alignment(horizontal="right", vertical="top")
        ws[f"F{total_row}"] = f"=SUM(F2:F{last_data_row})"
        ws.cell(row=total_row, column=6).number_format = IDR_FORMAT
        for col in range(1, 7):
            ws.cell(row=total_row, column=col).border = thin_border
            ws.cell(row=total_row, column=col).font = Font(bold=True)

        # PPN row
        ppn_row = total_row + 1
        ws[f"E{ppn_row}"] = "PPN (11%)"
        ws[f"E{ppn_row}"].alignment = Alignment(horizontal="right", vertical="top")
        ws[f"F{ppn_row}"] = f"=F{total_row}*0.11"
        ws.cell(row=ppn_row, column=6).number_format = IDR_FORMAT
        for col in range(1, 7):
            ws.cell(row=ppn_row, column=col).border = thin_border

        # Grand Total row
        grand_row = ppn_row + 1
        ws[f"E{grand_row}"] = "Grand Total"
        ws[f"E{grand_row}"].font = Font(bold=True)
        ws[f"E{grand_row}"].alignment = Alignment(horizontal="right", vertical="top")
        ws[f"F{grand_row}"] = f"=F{total_row}+F{ppn_row}"
        ws.cell(row=grand_row, column=6).number_format = IDR_FORMAT
        for col in range(1, 7):
            ws.cell(row=grand_row, column=col).border = thin_border
            ws.cell(row=grand_row, column=col).font = Font(bold=True)

        # Auto-fit column widths first (before height calculation uses them)
        for col in range(1, 7):
            max_length = 0
            column = get_column_letter(col)
            for row in range(1, grand_row + 1):
                cell = ws[f"{column}{row}"]
                # Skip if part of merged cell
                if cell.coordinate in [mc.coord for mc in ws.merged_cells.ranges]:
                    continue
                cell_value = cell.value
                if cell_value is not None:
                    length = len(str(cell_value)) * 1.3
                    if length > max_length:
                        max_length = length
            # Fix column A (NO) width
            if column == 'A':
                ws.column_dimensions[column].width = 6
            else:
                ws.column_dimensions[column].width = max(max_length + 2, 10)
            # URAIAN column (B) width - capped at 50
            if column == 'B':
                ws.column_dimensions[column].width = min((max_length + 2) * 2, 50)

        # Calculate grand total with PPN
        grand_total_value = sum(float(item.get("harga_satuan", 0)) * float(item.get("volume", 0)) for item in items)
        ppn_value = grand_total_value * 0.11
        total_with_ppn = grand_total_value + ppn_value

        # Terbilang merged cell spanning Total:Grand rows
        ws.merge_cells(f'A{total_row}:D{grand_row}')
        ws[f'A{total_row}'] = "Terbilang:\n" + number_to_terbilang(total_with_ppn)
        ws[f'A{total_row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Calculate merged width for wrap calculation (now columns are auto-fitted)
        merged_width = sum(ws.column_dimensions[get_column_letter(c)].width for c in range(1, 5))
        bilang_text = number_to_terbilang(total_with_ppn)
        bilang_lines = len(bilang_text) / merged_width
        single_row_height = int(max(20, bilang_lines * 15))
        ws.row_dimensions[total_row].height = single_row_height

        return ws

    def save_workbook(self, wb: Workbook, output_path: str) -> None:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    def _calculate_required_height(self, text: str, col_width: float, font_size: float = 11) -> float:
        """Calculate row height based on wrapped text"""
        # ~1 char = 0.5 * font_size in points at default zoom
        chars_per_line = col_width / (font_size * 0.55)
        lines = max(1, len(text) / chars_per_line)
        return max(20, lines * 15)