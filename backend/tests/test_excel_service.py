import pytest
import sys
from pathlib import Path

# Add backend to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from services.excel_service import ExcelService
from openpyxl import load_workbook
from pathlib import Path


def test_add_items_table_creates_correct_headers():
    """Headers should be: NO, URAIAN, VOLUME, SATUAN, HARGA SATUAN, JUMLAH HARGA"""
    service = ExcelService()
    wb = service.create_workbook()
    items = [
        {"uraian": "item 1", "volume": 2, "satuan": "pcs", "harga_satuan": 10000},
        {"uraian": "item 2", "volume": 1, "satuan": "lot", "harga_satuan": 50000},
    ]
    ws = service.add_items_table(wb, items)

    headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
    assert headers == ["NO", "URAIAN", "VOLUME", "SATUAN", "HARGA SATUAN", "JUMLAH HARGA"]


def test_jumlah_harga_formula():
    """JUMLAH HARGA column should contain =C*D formula"""
    service = ExcelService()
    wb = service.create_workbook()
    items = [
        {"uraian": "item 1", "volume": 2, "satuan": "pcs", "harga_satuan": 10000},
    ]
    ws = service.add_items_table(wb, items)

    # Row 2 = first data row
    assert ws["F2"].value == "=C2*E2"


def test_summary_rows_formulas():
    """Total, PPN, Grand Total rows with correct formulas"""
    service = ExcelService()
    wb = service.create_workbook()
    items = [
        {"uraian": "item 1", "volume": 2, "satuan": "pcs", "harga_satuan": 10000},
    ]
    ws = service.add_items_table(wb, items)

    last_data_row = 2  # 1 header + 1 item

    # Total row
    total_row = last_data_row + 1  # row 3
    assert ws[f"E{total_row}"].value == "Total"
    assert ws[f"F{total_row}"].value == f"=SUM(F2:F{last_data_row})"

    # PPN row
    ppn_row = total_row + 1  # row 4
    assert ws[f"E{ppn_row}"].value == "PPN (11%)"
    assert ws[f"F{ppn_row}"].value == f"=F{total_row}*0.11"

    # Grand Total row
    grand_row = ppn_row + 1  # row 5
    assert ws[f"E{grand_row}"].value == "Grand Total"
    assert ws[f"F{grand_row}"].value == f"=F{total_row}+F{ppn_row}"


def test_generate_output_file():
    """Generate actual XLSX file for manual inspection - edge cases"""
    items = [
        {"uraian": "Laptop Kerja", "volume": 10, "satuan": "Unit", "harga_satuan": 30000000},
        {"uraian": "Monitor LED 24 inch", "volume": 10, "satuan": "Unit", "harga_satuan": 5000000},
        {"uraian": "Keyboard Wireless", "volume": 10, "satuan": "Unit", "harga_satuan": 1000000},
        {"uraian": "Mouse Wireless", "volume": 10, "satuan": "Unit", "harga_satuan": 600000},
        {"uraian": "Headset Logitech", "volume": 10, "satuan": "Unit", "harga_satuan": 1600000},
        {"uraian": "Printer Laser", "volume": 5, "satuan": "Unit", "harga_satuan": 7000000},
        {"uraian": "UPS 1000VA", "volume": 5, "satuan": "Unit", "harga_satuan": 2400000},
        {"uraian": "Kabel LAN Cat6 100m", "volume": 20, "satuan": "Roll", "harga_satuan": 900000},
        {"uraian": "Switch Hub 24 Port", "volume": 3, "satuan": "Unit", "harga_satuan": 5000000},
        {"uraian": "Router WiFi Enterprise", "volume": 3, "satuan": "Unit", "harga_satuan": 10000000},
        {"uraian": "Rack Server 42U", "volume": 1, "satuan": "Unit", "harga_satuan": 50000000},
        {"uraian": "SSD 512GB", "volume": 20, "satuan": "Unit", "harga_satuan": 2400000},
        {"uraian": "HDD External 2TB", "volume": 10, "satuan": "Unit", "harga_satuan": 1800000},
        {"uraian": "Software Office Suite", "volume": 15, "satuan": "Lisensi", "harga_satuan": 6000000},
        {"uraian": "Maintenance Service 1 Tahun", "volume": 5, "satuan": "Paket", "harga_satuan": 40000000},
        {"uraian": "Pengadaan dan Instalasi Jaringan Fiber Optic Multi-Core Single-Mode 12 Core Untuk Area Data Center Gedung Utama Lantai 3-7 Termasuk Testing dan Commisioning", "volume": 2, "satuan": "Lot", "harga_satuan": 150000000},
    ]

    service = ExcelService()
    wb = service.create_workbook()
    service.add_items_table(wb, items)

    output_path = Path(__file__).parent.parent / "output_test"
    output_path.mkdir(exist_ok=True)
    service.save_workbook(wb, str(output_path / "RAB_test.xlsx"))
    print(f"\nSaved to {output_path / 'RAB_test.xlsx'}")